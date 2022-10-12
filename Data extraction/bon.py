from sqlite3 import Cursor
from xml.etree.ElementTree import Comment
import psycopg2
import csv
from pathlib import Path
import pandas as pd
import unicodedata
import os
import numpy as np
import psycopg2.extras as extras
from io import StringIO
import openpyxl


def create_tables():
    commands = (
        """DROP TABLE IF EXISTS nests_2021""",
        """DROP TABLE IF EXISTS nests_2022""",
        """ 
        create table nests_2021(
        nest_id varchar(255),
        place varchar(255),
        discovered timestamp,
        northing float(10),
        easting float(10),
        maps_link text,
        crop varchar(255)
        )
        """,
        """
        create table nests_2022(
        nest_id varchar(255),
        place varchar(255),
        discovered timestamp,
        northing float(10),
        easting float(10),
        maps_link text,
        crop varchar(255)
        )
        """
    )

    # do the statements
    conn = None
    try:
        # connect 
        conn = psycopg2.connect("dbname='c5dv202_vt21_bio17sla' user='c5dv202_vt21_bio17sla' host='postgres.cs.umu.se' password='gf9uYTcrPusW'")
        cur = conn.cursor()
        for command in commands:
            cur.execute(command)
        # close communication with the PostgreSQL database server
        cur.close()
        # commit the changes
        conn.commit()
    except(Exception, psycopg2.DatabaseError) as error:
        print(error)
    finally:
        if conn is not None:
            conn.close()


def insert_to_table(cursor, connection, file_name, query):
    data_folder = Path("Data/Metadata bon/")
    dataframe = openpyxl.load_workbook(data_folder / file_name)

    # Define variable to read sheet
    dataframe1 = dataframe.active
    # first row with data is row 2
    for i in range(2, dataframe1.max_row + 1):
        nest_id = dataframe1.cell(row = i, column = 1).value
        place = str(dataframe1.cell(row = i, column = 2).value)
        discovered = str(dataframe1.cell(row = i, column = 3).value)
        index = discovered.find('x')
        if index != -1:
            discovered = discovered.replace('x','1')
        index2 = discovered.find('2')
        discovered = discovered.replace(discovered[0:index2], '',1)
        dd_lat = dataframe1.cell(row = i, column = 4).value
        dd_long = dataframe1.cell(row = i, column = 5).value
        maps_link = ("https://www.google.com/maps/place/%s,%s" % (dd_lat, dd_long))
        crop = str(dataframe1.cell(row = i, column = 6).value)
        print(nest_id, place, discovered, dd_lat, dd_long)
        # when the data is ok, insert to sql table.
        #finally insert some stuff
        try:
            if str(nest_id) != "None":
                cursor.execute(query, (nest_id, place, discovered, dd_lat, dd_long, maps_link, crop))
                connection.commit()
        except(Exception, psycopg2.DatabaseError) as error:
            print("Error %s" % error)
            connection.rollback()
            cursor.close()
            return 1


if __name__ == '__main__':
    create_tables()
    # connect to database handle errors
    try:
        conn = psycopg2.connect("dbname='c5dv202_vt21_bio17sla' user='c5dv202_vt21_bio17sla' host='postgres.cs.umu.se' password='gf9uYTcrPusW'")
        # now that connected to database, define a cursor to be able to execute commands
        cur = conn.cursor()
        print("connected to database")

        # start inserting them querieeezz
        query = """INSERT INTO nests_2021 (nest_id, place, discovered, northing, easting, maps_link, crop)
                                VALUES (%s, %s, %s, %s, %s, %s, %s)"""

        insert_to_table(cur,conn,"Bon 2021.xlsx", query)

        query = """INSERT INTO nests_2022 (nest_id, place, discovered,northing, easting, maps_link, crop)
                                VALUES (%s, %s, %s, %s, %s, %s, %s)"""
        insert_to_table(cur,conn,"Bon 2022.xlsx", query)

        
    except:
        print("Unable to connect to the database, please check connection and start again")
        

