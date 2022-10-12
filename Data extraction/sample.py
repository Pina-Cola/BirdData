from pydoc import describe
import psycopg2
from pathlib import Path
import openpyxl
from datetime import date


def create_tables():
    commands = (
        """DROP TABLE IF EXISTS ringmarkningsdata""",
        """DROP TABLE IF EXISTS sample_data""",
        """ 
        
    create table ringmarkningsdata(
        year varchar(255),
        ring_id int,
        northing float(10),
        easting float(10),
        maps_link text
        )
        """,
        """
        create table sample_data(
        sample_id varchar(255),
        nest_id varchar(255),
        description varchar(255),
        northing float(10),
        easting float(10),
        maps_link text
        )
        """
    )
    
    # do the statements
    conn = None
    try:
        # connect 
        conn = psycopg2.connect("dbname='c5dv202_vt21_bio17sla' user='c5dv202_vt21_bio17sla' host='postgres.cs.umu.se' password='gf9uYTcrPusW'")
        cur = conn.cursor()
        for command in commands:
            cur.execute(command)
        # close communication with the PostgreSQL database server
        cur.close()
        # commit the changes
        conn.commit()
    except(Exception, psycopg2.DatabaseError) as error:
        print(error)
    finally:
        if conn is not None:
            conn.close()

def insert_into_ringmarkningsdata():
    dataframe = openpyxl.load_workbook("Data/Ringmärkning/Ringmärkningsdata.xlsx")

    # Define variable to read sheet
    dataframe1 = dataframe.active

    # connect:
    conn = None
    try:
        # connect 
        conn = psycopg2.connect("dbname='c5dv202_vt21_bio17sla' user='c5dv202_vt21_bio17sla' host='postgres.cs.umu.se' password='gf9uYTcrPusW'")
        cur = conn.cursor()
    except(Exception, psycopg2.DatabaseError) as error:
        print(error)
    

    query = """INSERT INTO  ringmarkningsdata(year, ring_id, northing, easting, maps_link)
                                VALUES (%s, %s, %s, %s, %s)"""

    for i in range(7,dataframe1.max_row + 1):
        year = dataframe1.cell(row = i, column = 1).value
        ring_id = dataframe1.cell(row = i, column = 2).value
        northing = dataframe1.cell(row = i, column = 3).value
        easting = dataframe1.cell(row = i, column = 4).value
        maps_url = ("https://www.google.com/maps/place/%s,%s" % (northing, easting))

        print(year,ring_id,northing,easting,maps_url)
        cur.execute(query, (year, ring_id, northing, easting, maps_url))

    
    # commit changes and close connections.
    try:
        # close communication with the PostgreSQL database server
        cur.close()
        # commit the changes
        conn.commit()
    except(Exception, psycopg2.DatabaseError) as error:
        print(error)
    finally:
        conn.close()


def insert_into_sampledata():
    dataframe = openpyxl.load_workbook("Data/Samples/Sample data 2020-2022.xlsx")

    # Define variable to read sheet
    dataframe1 = dataframe.active

    # connect:
    conn = None
    try:
        # connect 
        conn = psycopg2.connect("dbname='c5dv202_vt21_bio17sla' user='c5dv202_vt21_bio17sla' host='postgres.cs.umu.se' password='gf9uYTcrPusW'")
        cur = conn.cursor()
    except(Exception, psycopg2.DatabaseError) as error:
        print(error)

    query = """INSERT INTO  sample_data(sample_id, nest_id, description, northing, easting, maps_link)
                                VALUES (%s, %s, %s, %s, %s, %s)"""

    for i in range(14,dataframe1.max_row + 1):
        sample_id = dataframe1.cell(row = i, column = 1).value
        
        if str(dataframe1.cell(row = i, column = 2).value).find(' ') == -1:
            temp = str(dataframe1.cell(row = i, column = 2).value).split('2',1)
            nest_id = ' 2'.join(temp)
        else:
            nest_id = dataframe1.cell(row=i, column = 2).value

        if nest_id.find('-') != -1:
            nest_id = nest_id.replace('-', '/')
        # if nest_id contains - then replace with / + add a space between 2020
        description = dataframe1.cell(row = i, column = 3).value
        northing = dataframe1.cell(row = i, column = 4).value
        easting = dataframe1.cell(row = i, column = 5).value
        maps_url = ("https://www.google.com/maps/place/%s,%s" % (northing, easting))
        cur.execute(query, (sample_id, nest_id,description, northing, easting, maps_url))
        print(sample_id,nest_id,description,northing,easting,maps_url)



    # commit changes and close connections.
    try:
        # close communication with the PostgreSQL database server
        cur.close()
        # commit the changes
        conn.commit()
    except(Exception, psycopg2.DatabaseError) as error:
        print(error)
    finally:
        conn.close()


if __name__ == '__main__':
    create_tables()
    insert_into_ringmarkningsdata()
    insert_into_sampledata()
